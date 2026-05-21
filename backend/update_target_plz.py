"""Aktualisiert PLZ-Feld der Targets aus der Kunden-Excel."""
import os
import re
import sys
from openpyxl import load_workbook
from azure.data.tables import TableServiceClient

EXCEL_PATH = "/Users/annagiza-braun/Downloads/Kunden und Ex-Kunden.xlsx"
CONN = os.environ.get("AZURE_TABLE_STORAGE_CONNECTION_STRING")
if not CONN:
    sys.exit("AZURE_TABLE_STORAGE_CONNECTION_STRING fehlt")

print("Lade Excel…")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Sheet1"]
rows = ws.iter_rows(values_only=True)
header = next(rows)

# Spalten-Indizes
COL_PLZ = 8  # Kontakt: Postleitzahl
COL_ORT = 9  # Kontakt: Stadt
COL_FIRMA = 5  # Kontakt: Firmenname
COL_MBNR = 102 # Kontakt: ITUKV ProjNr

# mb-Nr -> {plz, ort, firma}
mb_to_data = {}
for row in rows:
    if not row or len(row) < 105:
        continue
    mbNr_raw = (row[COL_MBNR] or "")
    if not mbNr_raw:
        continue
    m = re.search(r"mb-?(\d+)", str(mbNr_raw), re.IGNORECASE)
    if not m:
        continue
    mbNr = f"mb-{m.group(1)}"
    plz = str(row[COL_PLZ]).strip() if row[COL_PLZ] else ""
    ort = str(row[COL_ORT]).strip() if row[COL_ORT] else ""
    firma = str(row[COL_FIRMA]).strip() if row[COL_FIRMA] else ""
    if plz:
        # Erste vollständige Daten gewinnen
        if mbNr not in mb_to_data:
            mb_to_data[mbNr] = {"plz": plz, "ort": ort, "firma": firma}

print(f"\nGefundene PLZ pro mb-Nummer:")
for mb, data in sorted(mb_to_data.items()):
    print(f"  {mb}: {data['plz']} {data['ort']} ({data['firma']})")

# Targets updaten
print("\nUpdate Targets…")
svc = TableServiceClient.from_connection_string(CONN)
tc = svc.get_table_client("targets")
updated = 0
for t in tc.list_entities():
    mbNr = t.get("mbNr", "")
    if mbNr in mb_to_data:
        data = mb_to_data[mbNr]
        if not t.get("plz"):
            t["plz"] = data["plz"]
        if not t.get("ort"):
            t["ort"] = data["ort"]
        if not t.get("firma"):
            t["firma"] = data["firma"]
        tc.update_entity(dict(t))
        updated += 1
        print(f"  ✓ {mbNr} → PLZ {data['plz']}")

print(f"\n{updated} Targets aktualisiert.")
