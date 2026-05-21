"""Aktualisiert PLZ-Feld der existierenden Kontakte aus der Excel."""
import os
import sys
from openpyxl import load_workbook
from azure.data.tables import TableServiceClient

EXCEL_PATH = "/Users/annagiza-braun/Downloads/Kunden und Ex-Kunden.xlsx"
CONN = os.environ.get("AZURE_TABLE_STORAGE_CONNECTION_STRING")
if not CONN:
    sys.exit("AZURE_TABLE_STORAGE_CONNECTION_STRING fehlt")

print("Lade Excel…")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Sheet1"]
rows = ws.iter_rows(values_only=True)
header = next(rows)

# Spalten finden
COL_EMAIL = 2     # Kontakt: E-Mail-Adresse
COL_PLZ = 8       # Kontakt: Postleitzahl

# Email -> PLZ Mapping aus Excel
email_to_plz = {}
for row in rows:
    if not row or len(row) < 10:
        continue
    email = (row[COL_EMAIL] or "").strip().lower() if row[COL_EMAIL] else ""
    plz = str(row[COL_PLZ]).strip() if row[COL_PLZ] else ""
    if email and plz:
        email_to_plz[email] = plz

print(f"PLZ-Mapping: {len(email_to_plz)} E-Mails")

# Kontakte updaten
svc = TableServiceClient.from_connection_string(CONN)
tc = svc.get_table_client("kontakte")

updated = 0
total = 0
for k in tc.list_entities():
    total += 1
    email = (k.get("email","") or "").lower().strip()
    if not email:
        continue
    plz = email_to_plz.get(email)
    if plz and not k.get("plz"):
        k["plz"] = plz
        tc.update_entity(dict(k))
        updated += 1
        if updated <= 5:
            print(f"  {email}: PLZ={plz}")

print(f"\n✓ {updated} von {total} Kontakten mit PLZ aktualisiert")
