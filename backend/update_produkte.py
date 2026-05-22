"""Importiert Produkt-Flags zu den bestehenden Kontakten aus der Kunden-Excel."""
import os
import sys
from openpyxl import load_workbook
from azure.data.tables import TableServiceClient

EXCEL_PATH = "/Users/annagiza-braun/Downloads/Kunden und Ex-Kunden.xlsx"
CONN = os.environ.get("AZURE_TABLE_STORAGE_CONNECTION_STRING")
if not CONN:
    sys.exit("AZURE_TABLE_STORAGE_CONNECTION_STRING fehlt")

COL_EMAIL = 2
# Produkt-Flags (Boolean): Spalten-Index -> Feld-Name
FLAGS = {
    'hatUC': 53,
    'hatUCS': 28,
    'hatMC': 62,
    'hatFKE': 64,
    'hatUVE': 95,
    'hatVME': 70,
    'hatKIwerkOne': 50,
    'hatMSQ': 76,
    'hatKMQ': 26,
    'hatKIT': 83,
    'hatKK': 106,
    'imITUKV': 101,
}

print("Lade Excel…")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Sheet1"]
rows = ws.iter_rows(values_only=True)
next(rows)  # header

# Email -> Produkt-Flags
email_to_flags = {}
for row in rows:
    if not row or len(row) < 110:
        continue
    email = (row[COL_EMAIL] or "")
    email = str(email).strip().lower() if email else ""
    if not email:
        continue
    flags = {}
    for field, idx in FLAGS.items():
        if idx < len(row):
            v = row[idx]
            # bool oder ja/nein-string
            if v is True or (isinstance(v, str) and v.lower() in ('ja', 'yes', 'true')):
                flags[field] = True
    if flags:
        email_to_flags[email] = flags

print(f"E-Mails mit Produkten: {len(email_to_flags)}")

# Verteilung
from collections import Counter
all_flags = Counter()
for flags in email_to_flags.values():
    for f in flags:
        all_flags[f] += 1
print("Verteilung:")
for f, n in all_flags.most_common():
    print(f"  {f}: {n}")

# Update Kontakte
print("\nUpdate Kontakte…")
svc = TableServiceClient.from_connection_string(CONN)
tc = svc.get_table_client("kontakte")
updated = 0
for k in tc.list_entities():
    email = (k.get("email","") or "").lower().strip()
    if email in email_to_flags:
        for field, val in email_to_flags[email].items():
            k[field] = val
        tc.update_entity(dict(k))
        updated += 1

print(f"\n{updated} Kontakte aktualisiert.")
