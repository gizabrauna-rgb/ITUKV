"""
Einmaliger Import der Kunden-Excel in die ITUKV-Datenbank.
- Zeilen mit mb-Nummer → Targets
- Zeilen mit Kundenstatus = "Kunde" → Kontakte (CRM)
"""
import os
import sys
import uuid
import json
from datetime import datetime
from openpyxl import load_workbook
from azure.data.tables import TableServiceClient, UpdateMode

EXCEL_PATH = "/Users/annagiza-braun/Downloads/Kunden und Ex-Kunden.xlsx"
CONN = os.environ.get("AZURE_TABLE_STORAGE_CONNECTION_STRING")
if not CONN:
    sys.exit("Umgebungsvariable AZURE_TABLE_STORAGE_CONNECTION_STRING nicht gesetzt.")

DEFAULT_CHECKLISTE = [
    {"id": "1", "label": "Unternehmensbewertung", "done": False},
    {"id": "2", "label": "Fragebogen Unternehmensbewertung", "done": False},
    {"id": "3", "label": "Exposé erstellt", "done": False},
    {"id": "4", "label": "Mandat unterschrieben", "done": False},
    {"id": "5", "label": "Element-Raum eröffnet", "done": False},
    {"id": "6", "label": "Target beworben", "done": False},
    {"id": "7", "label": "Eingehende NDAs geprüft", "done": False},
    {"id": "8", "label": "Alle Dokumente vollständig", "done": False},
]

now = datetime.utcnow().isoformat()
svc = TableServiceClient.from_connection_string(CONN)
svc.create_table_if_not_exists("targets")
svc.create_table_if_not_exists("kontakte")
tc_targets = svc.get_table_client("targets")
tc_kontakte = svc.get_table_client("kontakte")

print("Lade Excel…")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Sheet1"]
rows = ws.iter_rows(values_only=True)
header = next(rows)

def idx(name):
    for i, h in enumerate(header):
        if h and name.lower() in str(h).lower():
            return i
    return None

# Spalten-Indizes
COL_VORNAME = idx("Vorname")
COL_NACHNAME = idx("Nachname")
COL_EMAIL = idx("E-Mail-Adresse") if idx("E-Mail-Adresse") is not None else idx("E-Mail")
COL_TELEFON = idx("Telefonnummer") if idx("Telefonnummer") is not None else idx("Telefon")
COL_FIRMA = idx("Firmenname") if idx("Firmenname") is not None else idx("Firma")
COL_WEBSITE = idx("Website")
COL_PLZ = idx("Postleitzahl") if idx("Postleitzahl") is not None else idx("PLZ")
COL_STADT = idx("Stadt")
COL_LAND = idx("Land")
COL_KUNDENSTATUS = idx("Kundenstatus")
COL_KUNDENNR = idx("Kundennummer")
COL_MBNR = idx("ITUKV ProjNr")
COL_NOTIZEN = idx("Notizen Salesliste") if idx("Notizen Salesliste") is not None else idx("Notizen")
COL_BEMERKUNG_ITUKV = idx("Bemerkung ITUKV")
COL_LEAD_HERKUNFT = idx("Leadherkunft")
COL_SCHWERPUNKT = idx("Unternehmenschwerpunkt")
COL_BRANCHE_NICHT_IT = idx("Bei Branche Nicht-ITler")

# UVE-Felder (Verkäufer-Infos)
COL_UVE_WANN = idx("Wann verkaufen")
COL_UVE_PREIS = idx("Preisvorstellung")
COL_UVE_WERT = idx("Wert")
# Investor-Felder
COL_INV_REGION = idx("Region zukaufen") if idx("Region zukaufen") is not None else idx("Region")
COL_INV_MA = idx("MA-Anzahl") if idx("MA-Anzahl") is not None else idx("Mitarbeiter")
COL_INV_TX = idx("Transaktionsgröße") if idx("Transaktionsgröße") is not None else idx("Transaktion")

print(f"Indizes: Vorname={COL_VORNAME}, Email={COL_EMAIL}, Firma={COL_FIRMA}, mbNr={COL_MBNR}")

targets_created = 0
targets_seen_mbnr = set()
kontakte_created = 0
kontakte_updated = 0
kontakte_seen_email = set()

# Bestehende Kontakte einlesen (fuer Dedup beim Re-Import)
print("Lade bestehende Kontakte für Dedup…")
existing_by_email = {}
existing_by_firma = {}
for ek in tc_kontakte.list_entities():
    e_mail = (ek.get("email") or "").lower().strip()
    e_firma = (ek.get("firma") or "").lower().strip()
    e_rk = ek.get("RowKey")
    if e_mail:
        existing_by_email[e_mail] = e_rk
    if e_firma:
        existing_by_firma.setdefault(e_firma, e_rk)
print(f"  {len(existing_by_email)} bestehende Email-Keys, {len(existing_by_firma)} Firma-Keys")

for row_idx, row in enumerate(rows, start=2):
    if not row:
        continue
    def g(col):
        if col is None or col >= len(row):
            return ""
        v = row[col]
        return str(v).strip() if v is not None else ""

    vorname = g(COL_VORNAME)
    nachname = g(COL_NACHNAME)
    email = g(COL_EMAIL).lower()
    firma = g(COL_FIRMA)
    mbNr_raw = g(COL_MBNR)
    kundenstatus = g(COL_KUNDENSTATUS)
    notizen = g(COL_NOTIZEN)
    bemerkung_itukv = g(COL_BEMERKUNG_ITUKV)

    # Skip leere Zeilen
    if not (email or firma or mbNr_raw):
        continue

    name = f"{vorname} {nachname}".strip()
    plz = g(COL_PLZ)
    ort = g(COL_STADT)
    telefon = g(COL_TELEFON)
    website = g(COL_WEBSITE)
    herkunft = g(COL_LEAD_HERKUNFT)
    schwerpunkt = g(COL_SCHWERPUNKT)
    branche_nicht_it = g(COL_BRANCHE_NICHT_IT)
    # Branche aus Schwerpunkt ableiten (cleane Hauptkategorie statt Freitext)
    branche_aus_schwerpunkt = ""
    sw_low = schwerpunkt.lower()
    if "msp" in sw_low or "managed" in sw_low or "systemhaus" in sw_low:
        branche_aus_schwerpunkt = "IT-Systemhaus / MSP"
    elif "softwareentw" in sw_low:
        branche_aus_schwerpunkt = "Softwareentwicklung"
    elif "cloud" in sw_low:
        branche_aus_schwerpunkt = "Cloud Provider"
    elif "beratung" in sw_low or "consulting" in sw_low or "datenschutz" in sw_low or "security" in sw_low or "nis2" in sw_low:
        branche_aus_schwerpunkt = "IT-Beratung / Security"
    elif "nicht-itler" in sw_low or "nicht itler" in sw_low:
        branche_aus_schwerpunkt = branche_nicht_it or "Sonstige (Nicht-IT)"
    elif schwerpunkt:
        branche_aus_schwerpunkt = schwerpunkt

    # mb-Nummer extrahieren (kann "mb-317" oder nur "317" sein)
    mbNr = ""
    if mbNr_raw:
        import re
        m = re.search(r"mb-?(\d+)", mbNr_raw, re.IGNORECASE)
        if m:
            mbNr = f"mb-{m.group(1)}"
        elif mbNr_raw.isdigit():
            mbNr = f"mb-{mbNr_raw}"

    # === TARGETS ===
    if mbNr and mbNr not in targets_seen_mbnr:
        # Nur anlegen wenn auch sinnvolle Daten vorhanden
        if firma or name:
            targets_seen_mbnr.add(mbNr)
            tid = str(uuid.uuid4())
            entity = {
                "PartitionKey": "target", "RowKey": tid,
                "mbNr": mbNr,
                "verkaueferName": name or firma,
                "firma": firma,
                "email": email,
                "telefon": telefon,
                "website": website,
                "region": ort,
                "plz": plz,
                "branche": "IT-Systemhaus",
                "mitarbeiter": "",
                "umsatz": g(COL_UVE_WERT) or g(COL_UVE_PREIS),
                "beschreibung": (bemerkung_itukv + " " + notizen).strip(),
                "projekttyp": "Projekt Target",
                "status": "verfuegbar",
                "checklisteJson": json.dumps(DEFAULT_CHECKLISTE),
                "createdAt": now,
            }
            try:
                tc_targets.upsert_entity(entity)
                targets_created += 1
                print(f"  TARGET: {mbNr} – {entity['verkaueferName']} ({firma})")
            except Exception as e:
                print(f"  ERROR Target {mbNr}: {e}")

    # === KONTAKTE (CRM) ===
    # Regeln:
    # - Kunden / Ex-Kunden: weiterhin mit Email importiert (wie bisher)
    # - Nichtkunden: nur wenn Firma + PLZ + Ort vollstaendig
    # - Ausschluss-Statuses fliegen IMMER raus (auch wenn vollst. Adresse)
    AUSSCHLUSS_SUBSTRINGS = [
        "kunden-mitarbeiter", "nicht-itler", "fake", "presse",
        "konkurrenz", "dauerhaft", "kunden-dublette", "insolvent",
    ]
    ks_low = kundenstatus.lower()
    is_excluded = any(s in ks_low for s in AUSSCHLUSS_SUBSTRINGS)
    is_kunde_or_ex = ("kunde" in ks_low) and ("kunden-mitarbeiter" not in ks_low) and ("kunden-dublette" not in ks_low)
    adresse_vollstaendig = bool(firma and plz and ort)

    soll_importieren = False
    if is_excluded:
        soll_importieren = False
    elif is_kunde_or_ex and email:
        soll_importieren = True  # wie bisher
    elif adresse_vollstaendig:
        soll_importieren = True  # neue Regel: Nichtkunden mit kompletter Adresse

    # Dedup: erst Email, dann Firma (lower)
    dedup_key = email or f"firma::{firma.lower()}"
    if soll_importieren and dedup_key and dedup_key not in kontakte_seen_email:
        kontakte_seen_email.add(dedup_key)
        # Wenn Kontakt schon existiert (Re-Import), bestehende RowKey wiederverwenden
        existing_rk = None
        if email and email in existing_by_email:
            existing_rk = existing_by_email[email]
        elif firma.lower() in existing_by_firma:
            existing_rk = existing_by_firma[firma.lower()]
        kid = existing_rk or str(uuid.uuid4())
        is_update = existing_rk is not None

        # Typ + Such-/Biete-Texte
        typ = "Sonstige"
        sucht_text = ""
        bietet_text = ""
        if g(COL_INV_REGION) or g(COL_INV_TX):
            typ = "Strategisch"
            sucht_text = f"{g(COL_INV_REGION)} {g(COL_INV_MA)} {g(COL_INV_TX)}".strip()
        elif g(COL_UVE_WANN) or g(COL_UVE_PREIS):
            typ = "Verkäufer-Interesse"
            bietet_text = f"verkaufen: {g(COL_UVE_WANN)} {g(COL_UVE_PREIS)}".strip()

        # Nichtkunde-Klassifizierung
        ks_stripped = kundenstatus.strip()
        is_nichtkunde = (not ks_stripped) or (
            ks_stripped.lower() not in ("kunde", "ex-kunde") and
            "kunde" not in ks_stripped.lower() and
            "investor" not in ks_stripped.lower() and
            "partner" not in ks_stripped.lower()
        )
        # Effektiver Kundenstatus: leere Werte werden als "Nichtkunde" markiert,
        # damit Jenny im CRM-Tab filtern kann
        effektiver_status = ks_stripped or "Nichtkunde"

        entity = {
            "PartitionKey": "kontakt", "RowKey": kid,
            "name": name,
            "firma": firma,
            # Geschäftsführer: primärer Ansprechpartner (Vor + Nachname)
            "geschaeftsfuehrer": name,
            # Branche aus Schwerpunkt-Spalte ableiten, sonst leer
            "branche": branche_aus_schwerpunkt,
            "bietet": (bietet_text + (" · " + schwerpunkt if schwerpunkt else "")).strip(" ·"),
            "email": email,
            "telefon": telefon,
            "website": website,
            "plz": plz,
            "ort": ort,
            "typ": typ,
            "sucht": sucht_text,
            "kommentar": notizen,
            "herkunft": herkunft or "Excel-Import",
            "kundenstatus": effektiver_status,
            "kundennummer": g(COL_KUNDENNR),
            # Flags fuer einfaches Filtern in der UI
            "istKunde": ks_stripped.lower() == "kunde",
            "istExKunde": "ex-kunde" in ks_low,
            "istInvestor": "investor" in ks_low,
            "istNichtkunde": is_nichtkunde,
            # Match-Felder leer - Jenny pflegt nach
            "mitarbeiter": "", "umsatzTeur": "",
            "ebitMarge": "", "recurringPct": "",
            "createdAt": now,
            "updatedAt": now,
        }
        try:
            # Bei Update: nicht alle Felder ueberschreiben (createdAt erhalten,
            # ggf. manuell gepflegte Match-Felder nicht durch leere ersetzen)
            if is_update:
                entity.pop("createdAt", None)
                for f in ("mitarbeiter", "umsatzTeur", "ebitMarge", "recurringPct"):
                    if not entity.get(f):
                        entity.pop(f, None)
                tc_kontakte.update_entity(entity, mode=UpdateMode.MERGE)
                kontakte_updated += 1
            else:
                tc_kontakte.create_entity(entity)
                kontakte_created += 1
            if (kontakte_created + kontakte_updated) % 500 == 0:
                print(f"  … {kontakte_created + kontakte_updated} Kontakte verarbeitet")
        except Exception as e:
            print(f"  ERROR Kontakt {firma or email}: {e}")

print(f"\n✓ Fertig:")
print(f"  Targets angelegt: {targets_created}")
print(f"  Kontakte neu angelegt: {kontakte_created}")
print(f"  Kontakte aktualisiert: {kontakte_updated}")
